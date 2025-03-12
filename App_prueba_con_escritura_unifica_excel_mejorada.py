#!/usr/bin/env python3
#import speech_recognition as sr
#import subprocess

import numpy as np
import openpyxl
import os
from openpyxl import Workbook
#from openpyxl.utils import get_column_letter
from datetime import date

from scipy.interpolate import InterpolatedUnivariateSpline
from PySide6.QtWidgets import QFormLayout, QWidget, QLineEdit, QPushButton, QSpinBox, QDoubleSpinBox, QLabel, QApplication
from PySide6.QtWidgets import QButtonGroup, QCheckBox, QGridLayout, QMessageBox

potenciaKw=np.array([1,2,3,4,5,6,7,8,9,10])
Pack_1_6=np.array([60,25,16,10,6])
Pack_1_10=np.array([60,25,16,10,6])

contenido = os.listdir(os.getcwd())
if 'datos.xlsx' not in contenido:
    wb=Workbook()
    archivo_activo=wb.active
    archivo_activo.append(['Nombre','N° Local','UPS','Corriente [A]','Respaldo','Prueba 2 min', 'Mensaje', 'Fecha'])
    wb.save('datos.xlsx')

path='datos.xlsx'

wb = openpyxl.load_workbook(path)
archivo_activo = wb.active

class my_widget(QWidget):
    def __init__(self):
        super().__init__()
        #Label
        self.nombre_label= QLabel()
        self.nombre_label.setText("Nombre y número de local: ")
        self.numero_local=QSpinBox()
        self.numero_local.setMaximum(999)

        self.potencia_label= QLabel()
        self.potencia_label.setText("Potencia de UPS: ")
        self.P1=QCheckBox("6 kW")
        self.P2=QCheckBox("10 kW")

        self.UPS1=QCheckBox("UPS 1")
        self.UPS2=QCheckBox("UPS 2")
        self.UPS3=QCheckBox("UPS 3")

        self.P1.stateChanged.connect(lambda: self.deseleccionar_otros(self.P1))
        self.P2.stateChanged.connect(lambda: self.deseleccionar_otros(self.P2))

        self.UPS1.stateChanged.connect(lambda: self.deseleccionar_otros_UPS(self.UPS1))
        self.UPS2.stateChanged.connect(lambda: self.deseleccionar_otros_UPS(self.UPS2))
        self.UPS3.stateChanged.connect(lambda: self.deseleccionar_otros_UPS(self.UPS3))

        self.Corriente_consumida = QDoubleSpinBox()
        self.Corriente_consumida.setSuffix(" A")
        self.corriente_label= QLabel()
        self.corriente_label.setText("Corriente consumida: ")

        self.mensaje_label = QLabel()
        self.mensaje_label.setText("Mensaje: ")

        self.mensaje = QLabel()

        self.prueba_2_minutos_label = QLabel()
        self.prueba_2_minutos_label.setText("Con carga inicial completa. \nDespués de 2 minutos debe tener: ")
        self.prueba_2_minutos = QLabel()

        self.boton_calcular=QPushButton("Calcular")

        self.tiempo_de_respaldo = QLabel()
        self.tiempo_de_respaldo.setText("El respaldo mínimo esperado es: ")
        
        self.respaldo = QLabel()
        
        #LineEdit
        self.nombre_LE=QLineEdit()

        #Layout
        self.layout=QGridLayout()

        self.layout.addWidget(self.nombre_label,0,0)
        self.layout.addWidget(self.nombre_LE,0,1)
        self.layout.addWidget(self.numero_local,0,2)

        self.layout.addWidget(self.potencia_label,2,0)
        self.layout.addWidget(self.P1,2,1)
        self.layout.addWidget(self.P2,2,2)

        self.layout.addWidget(self.UPS1,1,0)
        self.layout.addWidget(self.UPS2,1,1)
        self.layout.addWidget(self.UPS3,1,2)

        self.layout.addWidget(self.corriente_label,3,0)
        self.layout.addWidget(self.Corriente_consumida,3,1)

        self.layout.addWidget(self.boton_calcular,5,1)


        self.layout.addWidget(self.tiempo_de_respaldo,6,0)
        self.layout.addWidget(self.respaldo,6,1)
        self.layout.addWidget(self.mensaje_label,8,0)
        self.layout.addWidget(self.mensaje,8,1)

        self.layout.addWidget(self.prueba_2_minutos_label,7,0)
        self.layout.addWidget(self.prueba_2_minutos,7,1)

        self.setLayout(self.layout)
        self.boton_calcular.clicked.connect(self.fcn_cal)

    def deseleccionar_otros(self, checkbox_actual):
        # Deseleccionar todos los demás checkboxes
        for checkbox in [self.P1, self.P2]:
            if checkbox != checkbox_actual:
                checkbox.setChecked(False)

    def deseleccionar_otros_UPS(self, checkbox_actual):
        # Deseleccionar todos los demás checkboxes
        for checkbox in [self.UPS1, self.UPS2, self.UPS3]:
            if checkbox != checkbox_actual:
                checkbox.setChecked(False)

    def fcn_cal(self):
        self.respaldo.clear()
#        respaldo=''
        self.mensaje.clear()
#        mensaje=''
        self.prueba_2_minutos.clear()
#        prueba2m=''

        nombre = str(self.nombre_LE.text())
#        n_local = str(self.numero_local.value())
        n_local=self.numero_local.value()
        Corriente = str(self.Corriente_consumida.value())+' A'

        if self.UPS1.isChecked():
            UPS = 1
        elif self.UPS2.isChecked():
            UPS= 2
        elif self.UPS3.isChecked():
            UPS= 3
        else:
            msg_box = QMessageBox()
            msg_box.setText("Seleccione UPS.")
            msg_box.exec()
            UPS=None

        carga=float()
        carga=float(220) * self.Corriente_consumida.value()
        if self.P1.isChecked():
            pack=Pack_1_6
            potencia=potenciaKw[0:len(Pack_1_6)]
        elif self.P2.isChecked():
            if self.Corriente_consumida.value() > 45:
                msgBox = QMessageBox()
                msgBox.setText("El sistema está sobrecargado.")
                msgBox.exec()
                #self.mensaje.setText("El sistema está sobrecargado, los cables están subdimensionados.")
            pack=Pack_1_10
            potencia=potenciaKw[0:len(Pack_1_10)]
#            amperaje_cable=45
        else:
            msg_box = QMessageBox()
            msg_box.setText("Debe seleccionar una potencia de UPS.")
            msg_box.exec()
#            self.mensaje.setText(str(self.mensaje.text())+"Debe seleccionar una Potencia de UPS. ")

        try:
            potencia=potencia*1000
#            self.prueba2m=0
            if carga > potencia[len(potencia)-1]:
                msg_box = QMessageBox()
                msg_box.setText("UPS sobrecargada.")
                msg_box.exec()
                self.mensaje.setText("La UPS está sobre cargada.")
                self.respaldo.setText("Sin respaldo adecuado.")
                self.prueba2m = 0   
            else :
                f=InterpolatedUnivariateSpline(potencia,pack)
                valor=f(carga)
                valor=round(float(valor))
                valor_esperado=round(float(valor*0.8))
                prueba_dos_min=(2*100)/valor_esperado
                self.respaldo.setText(str(valor_esperado)+" minutos")
                self.prueba_2_minutos.setText(str(round(100-prueba_dos_min))+" %")
                self.prueba2m = round(100-prueba_dos_min)
        except Exception as e:
            msg_box = QMessageBox()
            msg_box.setText(f"No se han podido realizar los cálculos. Compruebe los datos ingresados. Error: {e}")
            msg_box.exec()

        try:
            archivo_activo.append([nombre,n_local,UPS,Corriente,self.respaldo.text(),self.prueba2m,self.mensaje.text(),date.today().strftime("%d/%m/%Y")])
            wb.save('datos.xlsx')
        except Exception as e:
            msg_box = QMessageBox()
            msg_box.setText(f"No se han podido guardar los datos. Error: {e}")
            msg_box.exec()
        return

app=QApplication()
Mainwidget=my_widget()
Mainwidget.show()
app.exec()