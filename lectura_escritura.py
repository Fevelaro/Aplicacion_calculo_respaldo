import pandas as pd
import os
from os import getcwd
from openpyxl import Workbook, load_workbook

archivo = str(getcwd())+"\\excel_datos.xlsx"

wb=Workbook()
archivo_activo=wb.active
archivo_activo.append(['Nombre','N° Local','UPS','Corriente Consumida','Respaldo Mínimo','Prueba de dos minutos', 'Mensaje'])

def lector-escribidor(datos):
    try Workbook=load_workbook(archivo)



lector-escribidor():
    archivo = str(getcwd())+"\\excel_datos.xlsx"
    df = pd.read_excel(archivo, sheet_name='Sheet')
    
    df.
    return

lector-escribidor()